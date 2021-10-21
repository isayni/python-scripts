#!/usr/bin/env python
'''
    Dominik Cholewa 2021
    https://github.com/kikuchiyooo

    Script to automate making overviews and charts
    based on the built-in Windows monitoring solution
    using csv files
'''
import os
import sys
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter as letter
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Series, Reference

# === ARGUMENT MANAGEMENT ===
path = os.getcwd()
files = list()
for f in os.listdir(path):
    if os.path.isfile(f):
        f = path + '/' + f
        if f.endswith('.csv'):
            files.append(f)
if len(files) == 0:
    print('no csv files found')
    sys.exit()

CALC_MEMORY = False
for i in range(1, len(sys.argv)):
    if sys.argv[i] == '--used-mem':
        try:
            TOTAL_MEMORY = int(sys.argv[i+1])
        except:
            print('supply total memory size in MBs')
            sys.exit()
        CALC_MEMORY = True

# === DATA SET CLASS ===
class Dataset:
    def __init__(self):
        self.cols = list()
        self.chart = BarChart()

# === CREATING THE WORKBOOK ===
wb = Workbook()
ws = wb.active

# === VARIABLES ===
max_row = 1
datasets = list()

# === INSERTING DATA FROM THE CSV FILES ===
column = 1
for file in files:
    dataset = Dataset()
    # Windows PL encoding, also works with utf-8, so it's just in case
    with open(file, encoding='cp1250') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        line_count = 0; line_length = 1
        for line in csv_reader:
            line_count += 1
            if line_count == 1: # do just once but still in the loop
                line_length += len(line) - 1
                if CALC_MEMORY:
                    line_length+=1
                # defining the dataset's columns used (e.g ['A', 'B', 'C'])
                for i in range(column, column + line_length):
                    dataset.cols.append(letter(i))
            for i in range(line_length):
                # converting data fields to floats if possible
                try:
                    val = float(line[i])
                except:
                    try:
                        val = line[i]
                    except:
                        val = ''
                # insert each field of the line in a separate cell
                ws[letter(column + i) + str(line_count)] = val
        max_row = line_count if (line_count > max_row) else max_row
        # Changing the first cell to present the name of the file which the data is from
        ws[letter(column) + "1"].value = file.split('/')[-1].split('.csv')[0]
        column += line_length
    datasets.append(dataset)

# === CALCULATE USED MEMORY ===
if CALC_MEMORY:
    TOTAL_MEMORY_CELL = 'A' + str(max_row + len(datasets) + 2)
    ws[TOTAL_MEMORY_CELL].value = TOTAL_MEMORY
    for dataset in datasets:
        for c in dataset.cols:
            if 'Dostępna pamięć' in ws[c + '1'].value:
                free_mem_col = c
        col = dataset.cols[-1]

        ws[col + '1'].value = "Użyta pamięć (MB)"
        for i in range(2, max_row+1):
            ws[col + str(i)].value = f"=({TOTAL_MEMORY_CELL}-{free_mem_col}{i})"

# === STYLE & AVERAGES ===
# FIRST ROW
# filename in yellow and shorten the headers' names
series_row = str(max_row + 1)
avg_row = str(max_row + 2)

for dataset in datasets:
    #  STYLING
    # The yellow ID cell
    cell              = ws[dataset.cols[0] + "1"]
    cell.font         = Font(bold=True, color="000000")
    cell.fill         = PatternFill(fgColor="ffff00", fill_type="solid")

    bottom_cell       = ws['A' + avg_row]
    bottom_cell.value = cell.value
    bottom_cell.font  = Font(bold=True, color="000000")
    bottom_cell.fill  = PatternFill(fgColor="ffff00", fill_type="solid")
    # skipping the first column
    for c in dataset.cols[1:]:
        cell = ws[c + "1"]
        cell.fill = PatternFill(fgColor="339966", fill_type="solid")
        cell.font = Font(bold=True, color="ffffff")
        # Beautifying the header to just include the data category
        cell.value = cell.value.split('\\')[-1]

        if dataset == datasets[0]:
            bottom_cell = ws[c + series_row]
            bottom_cell.fill = PatternFill(fgColor="339966", fill_type="solid")
            bottom_cell.font = Font(bold=True, color="ffffff")
            # copying the value from top to bottom
            bottom_cell.value = cell.value

        # === AVERAGES  ===
    for i in range(2, len(dataset.cols) + 1):
        avg_cell = ws[letter(i) + avg_row]
        avg_cell.value = f"=AVERAGE({dataset.cols[i-1]}2:{dataset.cols[i-1]}{max_row})"
        avg_cell.fill = PatternFill(fgColor="e3d27d", fill_type="solid")
    avg_row = str(int(avg_row) + 1)

# === MAKING CHARTS ===
# chart = BarChart()
# chart.title = "Chart"
# chart.type = "col"

# === SAVING THE WORKBOOK ==
wb.save(path + "/monitoring.xlsx")
