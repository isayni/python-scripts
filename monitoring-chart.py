#!/usr/bin/env python
#
# Dominik Cholewa 2021
# https://github.com/kikuchiyooo
#
# Script to automate making overviews and charts
# based on the built-in Windows monitoring solution
# using csv files
#
import os
import sys
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter as letter
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Series, Reference

# === ARGUMENT MANAGEMENT ===
if len(sys.argv) != 2:
    exit()
path = sys.argv[1]
files = list()
if os.path.isdir(path):
    for f in os.listdir(path):
        if os.path.isfile(f):
            f = path + '/' + f
            if f.endswith('.csv'):
                files.append(f)
else:
    sys.exit()

# === DATA SET CLASS ===
class Dataset:
    def __init__(self):
        self.cols = list()
        self.chart = BarChart()

# === CREATING THE WORKBOOK ===
wb = Workbook()
ws = wb.active

# === VARIABLES ===
max_row = 1
datasets = list()

# === INSERTING DATA FROM THE CSV FILES ===
column = 1
for file in files:
    dataset = Dataset()
    # Windows PL encoding, also works with utf-8, so it's just in case
    with open(file, encoding='cp1250') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        line_count = 0; line_length = 1
        for line in csv_reader:
            line_count += 1
            if line_count == 1: # do just once but still in the loop
                line_length += len(line) - 1
                # defining the dataset's columns used (e.g ['A', 'B', 'C'])
                for i in range(column, column + line_length):
                    dataset.cols.append(letter(i))
            for i in range(line_length):
                # converting data fields to floats if possible
                try:
                    val = float(line[i])
                except:
                    val = line[i]
                # insert each field of the line in a separate cell
                ws[letter(column + i) + str(line_count)] = val
        max_row = line_count if (line_count > max_row) else max_row
        # Changing the first cell to present the name of the file which the data is from
        ws[letter(column) + "1"].value = file.split('/')[-1].split('.csv')[0]
        column += line_length
    datasets.append(dataset)

# === STYLE & AVERAGES ===
# FIRST ROW
# filename in yellow and shorten the headers' names
bottom_row = str(max_row + 1)
avg_row = str(max_row + 2)
for dataset in datasets:
    #  STYLING
    # The yellow ID cell
    cell              = ws[dataset.cols[0] + "1"]
    bottom_cell       = ws[dataset.cols[0] + bottom_row]
    cell.font         = Font(bold=True, color="000000")
    cell.fill         = PatternFill(fgColor="ffff00", fill_type="solid")
    bottom_cell.font  = Font(bold=True, color="000000")
    bottom_cell.fill  = PatternFill(fgColor="ffff00", fill_type="solid")
    bottom_cell.value = cell.value
    # skipping the first column
    for c in dataset.cols[1:]:
        cell        = ws[c + "1"]
        bottom_cell = ws[c + bottom_row]

        cell.fill        = PatternFill(fgColor="339966", fill_type="solid")
        bottom_cell.fill = PatternFill(fgColor="339966", fill_type="solid")
        cell.font        = Font(bold=True, color="ffffff")
        bottom_cell.font = Font(bold=True, color="ffffff")
        # Beautifying the header to just include the data category
        cell.value        = cell.value.split('\\')[-1]
        # copying the value from top to bottom
        bottom_cell.value = cell.value

        # === AVERAGES  ===
        avg_cell = ws[c + avg_row]

        avg_cell.value = f"=AVERAGE({c}2:{c}{max_row})"
        avg_cell.fill  = PatternFill(fgColor="e3d27d", fill_type="solid")

# === MAKING CHARTS ===
# chart = BarChart()
# chart.title = "Chart"
# chart.type = "col"

# === SAVING THE WORKBOOK ==
wb.save(path + "/monitoring.xlsx")
